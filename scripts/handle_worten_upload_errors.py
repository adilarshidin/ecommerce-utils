import os
import json
import pandas as pd
import re
from dotenv import load_dotenv
from mistralai import Mistral
from openpyxl import load_workbook

# =========================
# Config
# =========================
load_dotenv()

MODEL = "mistral-large-latest"
client = Mistral(api_key=os.getenv("MISTRAL_API_TOKEN"))

ERRORS_FILE = "input/worten_errors_bricolaje_y_construccion.xlsx"
PRODUCTS_FILE = "output/worten/bricolaje_y_construccion.xlsx"
OUTPUT_FILE = "output/worten/bricolaje_y_construccion_full.xlsx" 

# =========================
# Load Excel files
# =========================
errors_df = pd.read_excel(ERRORS_FILE)
products_df = pd.read_excel(PRODUCTS_FILE, sheet_name="Data", header=1)

# =========================
# Utils
# =========================
def extract_json(text: str) -> dict:
    text = text.strip()
    text = re.sub(r"```(?:json)?", "", text, flags=re.IGNORECASE).strip()
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if not match:
        return {}
    try:
        return json.loads(match.group())
    except json.JSONDecodeError:
        return {}

def get_image_urls(product_row: pd.Series) -> list[str]:
    image_cols = [c for c in product_row.index if c.startswith("image") and pd.notna(product_row[c])]
    return [product_row[c] for c in image_cols]

def extract_error_fields(error_text: str) -> set[str]:
    """
    Extracts field names from Worten error messages like:
    "2010|The attribute 'blade-length-cm' (...) must be an integer"
    """
    if not error_text:
        return set()
    return set(re.findall(r"'([^']+)'", str(error_text)))

# =========================
# Process each error
# =========================
results = []

for idx, err_row in errors_df.iterrows():
    product_id = err_row["product_id"]
    error_fields = extract_error_fields(err_row["errors"])
    missing_cols = list(error_fields)

    product_row = products_df[products_df["product_id"] == product_id]
    if product_row.empty:
        print(f"⚠ Product {product_id} not found in products file.")
        continue

    product_dict = product_row.iloc[0].to_dict()
    images = get_image_urls(product_row.iloc[0])
    product_dict["images"] = images

    instructions = (f"""
You are completing missing or INVALID product data for Worten marketplace.
Product data: {json.dumps(product_dict, ensure_ascii=False)}
Fields to fix: {missing_cols}

STRICT OUTPUT RULES:
- Return ONLY a JSON object.
- Keys MUST match the field names exactly.
- Values MUST be plain strings or numbers (NO objects, NO arrays).
- Do NOT include explanations or extra text.
- Fill ONLY the requested fields.

COLUMN-SPECIFIC RULES (MANDATORY):

- If 'product-dimensions' is requested:
    - Format: LxWxH cm (example: "100x50x50 cm")
    - Single string only

- If 'blade-length-cm' is requested:
    - INTEGER number only
    - No units, no decimals

- If 'safety-system_pt_PT' is requested:
    - Return EXACTLY ONE of:
    Sim | Não | Sí | Si | No | Não Aplicável | No aplicable | No Aplicable

- If 'product_name_pt_PT' or 'product_name_es_ES' is requested:
    - Return the product name with MAXIMUM of 150 characters.
""")

    message_content = [{"type": "text", "text": instructions}]
    for img_url in images:
        message_content.append({"type": "image_url", "image_url": img_url})

    try:
        response = client.chat.complete(
            model=MODEL,
            messages=[{"role": "user", "content": message_content}],
            temperature=0
        )

        raw = response.choices[0].message.content
        missing_values = extract_json(raw)
        print(raw)
        if not missing_values:
            print(f"⚠ No valid JSON returned for product {product_id}")
            continue

        if "mp_category" in missing_values:
            if not missing_values["mp_category"].startswith("Bricolaje y Construcción/"):
                print(f"⚠ Invalid category root for product {product_id}")
                continue

        if "blade-length-cm" in missing_values:
            if not isinstance(missing_values["blade-length-cm"], int):
                print(f"⚠ Invalid blade-length-cm for product {product_id}")
                continue

        if "safety-system_pt_PT" in missing_values:
            allowed = {
                "Sim", "Não", "Sí", "Si", "No",
                "Não Aplicável", "No aplicable", "No Aplicable"
            }
            if missing_values["safety-system_pt_PT"] not in allowed:
                print(f"⚠ Invalid safety-system_pt_PT for product {product_id}")
                continue

        output_row = {
            **product_dict,
            **missing_values,
            "__error_fields__": error_fields
        }
        results.append(output_row)

        print(f"✅ Processed product {product_id}")

    except Exception as e:
        print(f"❌ Error processing product {product_id}: {e}")
        continue

# =========================
# Save results to Excel preserving formatting & sheets
# =========================
if not results:
    print("⚠ No results to save.")
    exit(0)

# Load template workbook
wb = load_workbook(PRODUCTS_FILE)
if "Data" not in wb.sheetnames:
    raise ValueError("Sheet 'Data' not found in template")

ws = wb["Data"]

# Build column index from header row (row 2, same as your other script)
col_index = {}
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=2, column=col).value
    if header:
        col_index[header] = col

ANCHOR_COLUMN = "product_id"
if ANCHOR_COLUMN not in col_index:
    raise ValueError(f"Anchor column '{ANCHOR_COLUMN}' not found")

# Convert results to dict keyed by product_id for fast lookup
results_by_id = {
    str(r["product_id"]): r
    for r in results
    if "product_id" in r
}

updated = 0

# Iterate existing rows and fill only missing values
row = 3
while ws.cell(row=row, column=col_index[ANCHOR_COLUMN]).value:
    product_id = str(ws.cell(row=row, column=col_index[ANCHOR_COLUMN]).value)

    if product_id not in results_by_id:
        row += 1
        continue

    result = results_by_id[product_id]

    for field, value in result.items():
        if field not in col_index:
            continue

        cell = ws.cell(row=row, column=col_index[field])

        error_fields = result.get("__error_fields__", set())

        if field in error_fields:
            # Force rewrite for errored fields
            cell.value = value
        elif cell.value in (None, "", "nan"):
            # Normal first-pass fill
            cell.value = value

    updated += 1
    row += 1

# Save as new file
wb.save(OUTPUT_FILE)

print(f"✅ Completed! {updated} products enriched and saved to {OUTPUT_FILE}")
