import pandas as pd
import glob
import os
from openpyxl import load_workbook
from pathlib import Path

# =========================
# PATHS
# =========================
CSV_FILE = "output/all_listings_ready.csv"
XLSX_DIR = "templates/worten"
OUTPUT_DIR = "output/worten_filled"

Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

# =========================
# AMAZON → WORTEN MAPPING
# =========================
WORTEN_MAPPING = {
    "moda": {
        "APPAREL", "ONE_PIECE_OUTFIT", "COAT", "SHIRT", "PANTS", "SHORTS", "SKIRT",
        "SWEATER", "SWEATSHIRT", "VEST", "SOCKS", "UNDERPANTS", "BASE_LAYER_APPAREL_SET",
        "HAT", "SCARF", "EARMUFF", "SWIMWEAR", "SNOWSUIT", "APPAREL_GLOVES", "SPORT_ACTIVITY_GLOVE",
        "BELTS", "SUSPENDER", "HANDBAG", "HIP_FLASK", "WATCH", "SHOE_ACCESSORY",
        "APPAREL_BELT", "APPAREL_HEAD_NECK_COVERING", "BRACELET", "SMOKING_PIPE"
    },

    "ropa_y_calzado_deportivo": {
        "SHOES", "BOOT", "SPORT_HELMET", "AUTOMOTIVE_HELMET", "SPORTING_GOODS",
        "SPORT_EQUIPMENT_BAG_CASE", "KNEE_PAD", "SAFETY_GLASSES", "SNOWSHOE",
        "SPORT_BAT", "SHOE_BAG", "CELL_PHONE_HOLSTER", "WAIST_PACK"
    },

    "salud_bienestar_y_cuidados_para_bebe": {
        "FIRST_AID_KIT", "SELF_DEFENSE_SPRAY", "PROTECTIVE_GLOVE",
        "SAFETY_HARNESS", "WASTE_BAG", "PET_TOY"
    },

    "productos_de_cuidado_personal": {
        "COSMETIC_CASE", "BODY_PAINT", "HEALTH_PERSONAL_CARE",
        "CUIDADO_BUCAL", "CUIDADO_CARA", "CUIDADO_CABELLO"
    },

    "supermercado_bebidas_y_limpieza": {
        "FOOD", "CLEANING_AGENT", "PEST_CONTROL_DEVICE", "SOLID_FIRE_FUEL"
    },

    "muebles_y_accesorios": {
        "HOME_FURNITURE_AND_DECOR", "STOOL_SEATING", "TABLE",
        "STORAGE_BOX", "CADDY", "NETTING_COVER"
    },

    "deporte_aire_libre_y_viaje": {
        "TENT", "TARP", "HAMMOCK", "SLEEPING_BAG", "SLEEPING_MAT",
        "OUTDOOR_RECREATION_PRODUCT", "BACKPACK", "DUFFEL_BAG", "CARRIER_BAG_CASE",
        "HYDRATION_PACK", "AIR_MATTRESS", "ANCHOR_STAKE", "BICYCLE_LIGHT",
        "CAMPING_EQUIPMENT", "CARGO_STRAP", "NAVIGATION_COMPASS",
        "PORTABLE_STOVE", "SURVIVAL_KIT", "MESS_KIT",
        "AIR_GUN_PROJECTILE", "GUN_CLEANING_KIT", "GUN_HOLSTER", "GUN_SLING",
        "BINOCULAR", "WEAPON_CASE"
    },

    "bricolaje_y_construccion": {
        "AXE", "KNIFE", "MULTITOOL", "SCREW_GUN", "SHOVEL_SPADE", "SAW",
        "ADHESIVE_TAPES", "BUCKET", "CABLE_TIE", "CORD_ROPE",
        "ELASTIC_BAND", "FABRIC_APPLIQUE_PATCH", "KNIFE_SHARPENER", "SEWING_TOOL_SET",
        "CARABINER", "LOCK", "PAINT"
    },

    "hogar": {
        "COOKING_POT", "KITCHEN_KNIFE", "FLATWARE", "DISHWARE_PLACE_SETTING", "DISHWARE_PLATE",
        "THERMOS", "DRINKING_CUP", "BOTTLE", "TOWEL", "PILLOW", "BLANKET", "STORAGE_BAG",
        "BLADED_FOOD_PEELER", "CAN_OPENER", "KITCHEN", "MANUAL_FOOD_MILL_GRINDER", "BUCKLE"
    },

    "merchandising_&_gifting": {
        "GLITTER", "CHARM", "BADGE_HOLDER", "BANNER", "KEYCHAIN", "LABEL"
    },

    "smart_home": {
        "LIGHT_BULB", "LIGHT_FIXTURE", "HOME_LIGHTING_AND_LAMPS",
        "FLASHLIGHT", "UTILITY_HOLSTER_POUCH"
    },

    "fotografia_y_video": {
        "ACCESSORIOS_FOTOGRAFIA_Y_VIDEO",
        "CAMERAS", "OBJETIVOS_Y_FLASHES", "VIDEO",
        "CAMERA_FILM"
    },

    "mascotas": {
        "ACCESORIOS_PARA_ANIMALES", "COMIDA_PARA_ANIMALES",
        "HIGIENE_CUIDADO_Y_SALUD_PARA_ANIMALES", "JUGUETES_PARA_ANIMALES",
        "LOCALIZADORES_Y_SEGURIDAD_DE_ANIMALES", "MUEBLES_PARA_ANIMALES",
        "TRANSPORTE_DE_ANIMALES"
    },

    "electrodomesticos": {
        "AIRE_ACONDICIONADO_Y_CALOR", "CAFETERAS", "COCINAS",
        "CONGELADORES", "EQUIPOS_INDUSTRIALES", "FRIGORIFICOS_Y_NEVERAS",
        "INTEGRABLES", "LAVADORAS", "LAVAVAJILLAS", "LIMPIEZA_DE_SUPERFICIES",
        "MICROONDAS_Y_MINI_HORNOS", "PEQUENOS_ELECTRODOMESTICOS", "TRATAMIENTO_DE_ROPA",
        "VINOTECAS", "SPACE_HEATER"
    },

    "equipamiento_y_piezas_de_vehiculos": {
        "MOTOR_ENGINE_FUEL_TANK"
    },

    "musica": {
        "BRASS_AND_WOODWIND_INSTRUMENTS"
    },

    "libros_y_audiolibros": {
        "BLANK_BOOK"
    }
}

# =========================
# LOAD CSV
# =========================
df = pd.read_csv(CSV_FILE, dtype=str)

required_cols = {"seller-sku"}
missing = required_cols - set(df.columns)
if missing:
    raise ValueError(f"Missing required CSV columns: {missing}")

# Keep track of unmatched SKUs
all_skus = set(df["seller-sku"].dropna())
matched_skus_total = set()

# =========================
# PROCESS EACH XLSX
# =========================
xlsx_files = glob.glob(os.path.join(XLSX_DIR, "*.xlsx"))

total_skus_written = 0
# =========================
# WRITE SKUS + ADDITIONAL COLUMNS
# =========================

# Columns mapping: CSV column → XLSX columns
COLUMN_MAPPING = {
    "item-name": ["product_name_pt_PT", "product_name_es_ES", "product_description_pt_PT",
                  "product_description_es_ES"],
    "seller-sku": ["ean"],
    "amazon_product_type_es": ["type_pt_PT", "type_es_ES"],
    "manufacturer": ["product-brand"]
}

# Max number of images to write
MAX_IMAGES = 12

for xlsx_path in xlsx_files:
    filename = os.path.splitext(os.path.basename(xlsx_path))[0]

    if filename not in WORTEN_MAPPING:
        print(f"⏭️ Skipping {filename}.xlsx (no mapping)")
        continue

    amazon_types = WORTEN_MAPPING[filename]

    matched_df = df[df["amazon_product_type"].isin(amazon_types)].copy()
    if matched_df.empty:
        print(f"⚠️ No matches for {filename}.xlsx")
        continue

    matched_skus_total.update(matched_df["seller-sku"].dropna().tolist())
    total_skus_written += len(matched_df)
    print(f"📦 Writing {len(matched_df)} SKUs → {filename}.xlsx")

    wb = load_workbook(xlsx_path)
    if "Data" not in wb.sheetnames:
        print(f"❌ Sheet 'Data' not found in {filename}.xlsx")
        continue

    ws = wb["Data"]

    # Build column index mapping from header row 2
    col_index = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=2, column=col).value
        if header:
            col_index[header] = col

    # Ensure required XLSX columns exist
    for csv_col, xlsx_cols in COLUMN_MAPPING.items():
        for xlsx_col in xlsx_cols:
            if xlsx_col not in col_index:
                raise ValueError(f"'{xlsx_col}' column not found in {filename}.xlsx")

    # Find first empty row (starting from row 3)
    row = 3
    while ws.cell(row=row, column=col_index["product_id"]).value:
        row += 1

    # Write data row by row
    for _, row_data in matched_df.iterrows():
        # product_id (seller-sku)
        ws.cell(row=row, column=col_index["product_id"], value=row_data["seller-sku"])

        # CSV → XLSX columns mapping
        for csv_col, xlsx_cols in COLUMN_MAPPING.items():
            for xlsx_col in xlsx_cols:
                ws.cell(row=row, column=col_index[xlsx_col], value=row_data.get(csv_col, ""))

        # Images (image1..image12)
        for i in range(1, MAX_IMAGES + 1):
            csv_image_col = f"image{i}"
            xlsx_image_col = f"image{i}"
            if csv_image_col in row_data and xlsx_image_col in col_index:
                ws.cell(row=row, column=col_index[xlsx_image_col], value=row_data[csv_image_col])

        row += 1

    # Save to output directory
    output_path = os.path.join(OUTPUT_DIR, os.path.basename(xlsx_path))
    wb.save(output_path)

# =========================
# REPORT UNMATCHED SKUS
# =========================
unmatched_skus = all_skus - matched_skus_total
if unmatched_skus:
    print(f"⚠️ {len(unmatched_skus)} SKUs were not processed:")
    for sku in sorted(unmatched_skus):
        print(f" - {sku}")

print(f"✅ All applicable Worten sheets updated successfully in {OUTPUT_DIR}.")
print(f"Total SKUs written: {total_skus_written}")
