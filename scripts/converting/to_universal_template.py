import json
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def normalize_worten_key(key: str) -> str:
    return key.replace("_", " ").title()

def flatten_shopify(categories, parent_path=""):
    paths = []

    for cat in categories:
        current_path = (
            f"{parent_path} > {cat['name']}"
            if parent_path
            else cat["name"]
        )

        if "children" not in cat or not cat["children"]:
            paths.append(current_path)
        else:
            paths.extend(flatten_shopify(cat["children"], current_path))

    return paths

def flatten_worten(data, parent_path=""):
    paths = []

    if isinstance(data, dict):
        for key, value in data.items():
            normalized_key = normalize_worten_key(key)

            current_path = (
                f"{parent_path} > {normalized_key}"
                if parent_path
                else normalized_key
            )

            paths.extend(flatten_worten(value, current_path))

    elif isinstance(data, list):
        for item in data:
            paths.append(f"{parent_path} > {item}")

    return paths


# Load JSON
with open("templates/shopify_categories.json", "r", encoding="utf-8") as f:
    shopify_categories_data = json.load(f)
with open("templates/worten/product_categories.json", "r", encoding="utf-8") as f:
    worten_categories_data = json.load(f)

shopify_categories = flatten_shopify(shopify_categories_data)
worten_categories = flatten_worten(worten_categories_data)

wb = load_workbook("templates/universal_template.xlsx")

product_sheet = wb.active  # or wb["Products"]

# Create / reuse hidden sheet
shopify_sheet_name = "Shopify_Categories"
if shopify_sheet_name in wb.sheetnames:
    shopify_categories_sheet = wb[shopify_sheet_name]
else:
    shopify_categories_sheet = wb.create_sheet(shopify_sheet_name)

worten_sheet_name = "Worten_Categories"
if worten_sheet_name in wb.sheetnames:
    worten_categories_sheet = wb[worten_sheet_name]
else:
    worten_categories_sheet = wb.create_sheet(worten_sheet_name)

# Write categories to column A
for i, path in enumerate(shopify_categories, start=1):
    shopify_categories_sheet.cell(row=i, column=1, value=path)
for i, path in enumerate(worten_categories, start=1):
    worten_categories_sheet.cell(row=i, column=1, value=path)

shopify_categories_sheet.sheet_state = "hidden"
worten_categories_sheet.sheet_state = "hidden"

# Create dropdown
shopify_max_row = len(shopify_categories)
shopify_dv = DataValidation(
    type="list",
    formula1=f"='{shopify_sheet_name}'!$A$1:$A${shopify_max_row}",
    allow_blank=True
)

worten_max_row = len(worten_categories)
worten_dv = DataValidation(
    type="list",
    formula1=f"='{worten_sheet_name}'!$A$1:$A${worten_max_row}",
    allow_blank=True
)

product_sheet.add_data_validation(shopify_dv)
product_sheet.add_data_validation(worten_dv)

# Apply dropdown to G3 and below
shopify_dv.add("G3:G10000")
worten_dv.add("H3:H10000")

wb.save("templates/universal_template.xlsx")
