import re
import pandas as pd
import json
import os
from dotenv import load_dotenv
from mistralai import Mistral
from openpyxl import load_workbook
from pathlib import Path

# ---------------- CONFIG ----------------
excel_path = "templates/konus.xlsm"
sheet_name = "Plantilla"
csv_path = "input/konus_catalog.csv"
output_path = "output/amazon_konus.xlsm"
start_row = 6

# ---------------- ENV ----------------
load_dotenv()

MISTRAL_API_TOKEN = os.getenv("MISTRAL_API_TOKEN")
if not MISTRAL_API_TOKEN:
    raise RuntimeError("MISTRAL_API_TOKEN not found in environment")

# ---------------- LOAD FILES ----------------
Path(output_path).parent.mkdir(parents=True, exist_ok=True)
wb = load_workbook(excel_path, keep_vba=True)
ws = wb[sheet_name]

df = pd.read_csv(
    csv_path,
    sep=";",
    encoding="latin-1",
    engine="python"
)

# ---------------- AMAZON HEADERS ----------------
amazon_headers = [cell.value for cell in ws[4]]

# ---------------- CLEAN TEMPLATE ----------------
max_row = ws.max_row
if max_row >= start_row:
    ws.delete_rows(start_row, max_row - start_row + 1)

# ---------------- HELPERS ----------------
def clean_price(price_str):
    if not price_str:
        return None
    return re.sub(r"[^\d.]", "", str(price_str))


def clean_json(text):
    match = re.search(r"\{.*\}", text, re.S)
    if not match:
        raise ValueError(f"Invalid JSON from LLM: {text}")
    return json.loads(match.group())

def safe_dim(dims, key, subkey):
    """
    Safely extract nested dimension values from LLM output.
    Returns None if the structure is missing or null.
    """
    val = dims.get(key)
    if isinstance(val, dict):
        return val.get(subkey)
    return None

def complete_dim(dims, key):
    """
    Returns (value, unit) only if BOTH are present and non-null.
    Otherwise returns (None, None).
    """
    val = dims.get(key)
    if not isinstance(val, dict):
        return None, None

    value = val.get("value")
    unit = val.get("unit")

    if value is None or unit is None:
        return None, None

    return value, unit

# ---------------- DIRECT MAP ----------------
def direct_map(csv_row, enrichment):
    weight = csv_row.get("PesoNeto")
    if isinstance(weight, str) and weight.strip():
        weight = weight.lower()
        is_kg = "kg" in weight

        weight = (
            weight
            .replace("kg.", "")
            .replace("kg", "")
            .replace("gr.", "")
            .replace("gr", "")
            .replace("r", "")
            .replace(",", ".")
            .strip()
        )

        weight = float(weight)

        if is_kg:
            weight *= 1000

    medidas_raw = csv_row.get("Medidas")
    medidas = None
    if isinstance(medidas_raw, str):
        medidas_clean = medidas_raw.lower().replace("cm", "").replace(" ", "")
        medidas = "; ".join(medidas_clean.split("x"))

    dims = enrichment.get("dimensions", {})

    browse_node = None
    size = None
    product_type = enrichment.get("product_type")
    match product_type:
        case "RANGEFINDER":
            browse_node = "Bricolaje y herramientas > Herramientas manuales y eléctricas > Herramientas de medición y diseño > Herramientas para medición láser y accesorios > Telémetros láser (3053092031)"
        case "CAMERA_TRIPOD":
            browse_node = "Electrónica > Comunicación móvil y accesorios > Accesorios > Accesorios de foto y vídeo > Trípodes (21529596031)"
        case "MICROSCOPES":
            browse_node = "Industria, empresas y ciencia > Artículos educativos > Recursos para planes de estudios > Ciencia > Microscopios > Microscopios monoculares (1443222031)"
        case "AIMING_SCOPE_SIGHT":
            browse_node = "Electrónica > Fotografía y videocámaras > Prismáticos, telescopios y óptica > Dispositivos de visión nocturna (930881031)"
        case "MAGNIFIER":
            browse_node = "Electrónica > Comunicación móvil y accesorios > Accesorios > Accesorios de juegos > Ampliadores y magnificadores de pantalla (21529576031)"
        case "TELESCOPE":
            browse_node = "Electrónica > Fotografía y videocámaras > Prismáticos, telescopios y óptica > Monoculares (930884031)"
        case "BINOCULAR":
            browse_node = "Electrónica > Fotografía y videocámaras > Prismáticos, telescopios y óptica > Prismáticos (930885031)"
        case "FLASHLIGHT":
            browse_node = "Bricolaje y herramientas > Ferretería > Linternas y faroles de mano > Linternas (3053011031)"
            size = "pequeño"
        case "NAVIGATION_COMPASS":
            browse_node = "Deportes y aire libre > Electrónica y dispositivos > Brújulas (2928776031)"

    thickness_v, thickness_u = complete_dim(dims, "thickness")
    height_v, height_u = complete_dim(dims, "height")
    width_v, width_u = complete_dim(dims, "width")
    length_v, length_u = complete_dim(dims, "length")
    min_fd_v, min_fd_u = complete_dim(dims, "min_focal_distance")
    pkg_len_v, pkg_len_u = complete_dim(dims, "package_length")
    pkg_w_v, pkg_w_u = complete_dim(dims, "package_width")
    pkg_h_v, pkg_h_u = complete_dim(dims, "package_height")
    pkg_weight_v, pkg_weight_u = complete_dim(dims, "package_weight")

    formatted_entry = {
        "SKU": csv_row.get("EAN"),
        "SKU principal": csv_row.get("EAN"),
        "ID del producto": csv_row.get("EAN"),
        "Marca": csv_row.get("Marca"),
        "Fabricante": "Konus",
        "Nombre Modelo": csv_row.get("Modelo"),
        "Numero de modelo": enrichment.get("model_number"),
        "Nombre del producto": csv_row.get("Título_producto"),
        "Palabra clave genérica": csv_row.get("Descripción_corta"),
        "Descripción del producto": csv_row.get("Descripción_larga"),
        "Viñeta": enrichment.get("bullet"),
        "Nodos recomendados de búsqueda": browse_node,
        "Tamaño": size,
        "Tipo de producto": product_type,
        "Precio de venta recomendado (PVPR)": clean_price(csv_row.get("PVP FINAL")),
        "Tu precio EUR (Vender en Amazon, ES)": clean_price(csv_row.get("PVP FINAL")),
        "Precio de venta. EUR (Vender en Amazon, ES)": clean_price(csv_row.get("PVP FINAL")),
        "Estado del producto": "Nuevo",
        "Tipo de identificador del producto": "EAN",
        "Grupo de la marina mercante (ES)": "Nueva plantilla Envios",
        "Cumplimiento de código de canal (ES)": "DEFAULT",
        "Cantidad (ES)": "1",
        "Número de Artículos": "1",
        "Número de cajas": "1",
        "Componentes Incluidos": "1 artículo",
        "Numero de pieza": enrichment.get("part_number"),
        "Peso Artículo": weight,
        "Unidad de peso del artículo": "Gramos",
        "Tamaño del anillo": medidas,
        "Aumento máximo": dims.get("max_magnification"),
        "Distancia focal mínima": min_fd_v,
        "Longitud Paquete": pkg_len_v,
        "Unidad de longitud del paquete": pkg_len_u,
        "Ancho Paquete": pkg_w_v,
        "Unidad de anchura del paquete": pkg_w_u,
        "Altura Paquete": pkg_h_v,
        "Unidad de altura del paquete": "Centímetros",
        "Peso del paquete": pkg_weight_v,
        "Unidad del peso del paquete": "Kilogramos",
        "Garantía de Producto": "2",
        "¿Se necesitan baterías?": "No",
        "Normativas sobre mercancías peligrosas": "No aplicable",
        "Riesgo del GDPR": "No hay información electrónica almacenada.",
        "URL de la imagen principal": csv_row.get("Imagen_grande"),
        "País de origen": "Italia",
        "Color": "negro",
        "Mapa de color": "negro",
    }
    match product_type:
        case "FLASHLIGHT":
            formatted_entry["Fuente Alimentación"] = "Batería"
            formatted_entry["Etiquetado Eficiencia Energética UE"] = "A to G"
            formatted_entry["Eficiencia"] = "A"
            formatted_entry["Conteo de unidades"] = "1"
            formatted_entry["Tipo de conteo de unidades"] = "unidad"
            formatted_entry["¿Es frágil?"] = "No"
            formatted_entry["Unidad de la altura"] = "Centímetros"
            formatted_entry["Unidad de la longitud"] = "Centímetros"
            formatted_entry["Unidad del ancho"] = "Centímetros"
            formatted_entry["Unidad de altura del artículo"] = "Centímetros"
            formatted_entry["Unidad de grosor del artículo"] = "Centímetros"
            formatted_entry["Unidad del ancho del artículo"] = "Centímetros"
        case "MAGNIFIER":
            formatted_entry["¿Es frágil?"] = "No"
            formatted_entry["Tamaño"] = "pequeño"
        case "CAMERA_TRIPOD":
            formatted_entry["Material"] = "Plástico"
        case "NAVIGATION_COMPASS":
            formatted_entry["Material"] = "Plástico"
            formatted_entry["Seguridad Juguetes Edad EU Advertencia"] = "Ninguna advertencia aplicable"
            formatted_entry["Advertencia No Requisito Edad EU DSJ"] = "Ninguna advertencia aplicable"
        case "RANGEFINDER":
            formatted_entry["Material"] = "Plástico"
            formatted_entry["Tamaño"] = "pequeño"
            formatted_entry["Seguridad Juguetes Edad EU Advertencia"] = "Ninguna advertencia aplicable"
            formatted_entry["Advertencia No Requisito Edad EU DSJ"] = "Ninguna advertencia aplicable"
        case "AIMING_SCOPE_SIGHT":
            formatted_entry["Material"] = "Plástico"
            formatted_entry["Tamaño"] = "pequeño"
            if formatted_entry.get("Peso Artículo"):
                formatted_entry["Peso Artículo"] = round(
                    formatted_entry["Peso Artículo"] / 453.6, 2
                )
                formatted_entry["Unidad de peso del artículo"] = "Libras"
            formatted_entry["Peso Artículo Unidad"] = "Libras"
            formatted_entry["Seguridad Juguetes Edad EU Advertencia"] = "Ninguna advertencia aplicable"
            formatted_entry["Nombre del departamento"] = "Adultos unisex"
            formatted_entry["Advertencia No Requisito Edad EU DSJ"] = "Ninguna advertencia aplicable"
        case "TELESCOPE":
            formatted_entry["Unidad de grosor del artículo"] = "Centímetros"
            formatted_entry["Unidad de altura del artículo"] = "Centímetros"
            formatted_entry["Unidad del ancho del artículo"] = "Centímetros"

    return formatted_entry

# ---------------- LLM ----------------
mistral = Mistral(api_key=MISTRAL_API_TOKEN)

ALLOWED_PRODUCT_TYPES = [
    "NAVIGATION_COMPASS",
    "FLASHLIGHT",
    "BINOCULAR",
    "TELESCOPE",
    "MAGNIFIER",
    "AIMING_SCOPE_SIGHT",
    "MICROSCOPES",
    "CAMERA_TRIPOD",
    "RANGEFINDER",
]


def classify_product_enrichment(csv_row, mistral):
    prompt = f"""
You are enriching Amazon product listings.

Rules:
- Choose EXACTLY ONE product type from the allowed list
- Generate EXACTLY ONE factual bullet point
- Infer values only if clearly implied, otherwise return null
- Units must be metric
- Return ONLY valid JSON

Allowed product types:
{ALLOWED_PRODUCT_TYPES}
Monoculars are to be categorized as TELESCOPE

Product data:
- Title: {csv_row.get("Título_producto")}
- Short description: {csv_row.get("Descripción_corta")}
- Long description: {csv_row.get("Descripción_larga")}
- Family: {csv_row.get("Familia")}
- Model: {csv_row.get("Modelo")}

Return format:
{{
  "product_type": "ONE_OF_THE_ALLOWED_VALUES",
  "bullet": "Short factual bullet",
  "model_number": "Number of the model or the name of the product if model can not be deduced",
  "part_number": "Can be product name but below 40 characters",
  "dimensions": {{
    "thickness": {{ "value": number|null, "unit": "cm"|null }},
    "height": {{ "value": number|null, "unit": "cm"|null }},
    "width": {{ "value": number|null, "unit": "cm"|null }},
    "length": {{ "value": number|null, "unit": "cm"|null }},
    "max_magnification": number|null,
    "min_focal_distance": {{ "value": number|null, "unit": "cm"|null }},
    "package_length": {{ "value": number|null, "unit": "cm"|null }},
    "package_width": {{ "value": number|null, "unit": "cm"|null }},
    "package_height": {{ "value": number|null, "unit": "cm"|null }},
    "package_weight": {{ "value": number|null, "unit": "kg"|null }}
  }}
}}
"""

    res = mistral.chat.complete(
        model="mistral-small-latest",
        messages=[{"role": "user", "content": prompt}],
        stream=False,
    )

    data = clean_json(res.choices[0].message.content)

    if data.get("product_type") not in ALLOWED_PRODUCT_TYPES:
        raise ValueError(f"Invalid product type: {data.get('product_type')}")

    return data


# ---------------- PROCESS ----------------
current_row = start_row

for idx, csv_row in df.iterrows():
    csv_dict = csv_row.to_dict()
    sku = csv_dict.get("EAN")

    try:
        enrichment = classify_product_enrichment(csv_dict, mistral)
    except Exception as e:
        enrichment = {
            "product_type": None,
            "bullet": None,
            "warranty": None,
            "dimensions": {}
        }

    mapped = direct_map(csv_dict, enrichment)

    for col_idx, header in enumerate(amazon_headers, start=1):
        value = mapped.get(header)
        if value is not None:
            ws.cell(row=current_row, column=col_idx, value=value)

    wb.save(output_path)
    current_row += 1

print(f"Amazon XLSM generated: {output_path}")
