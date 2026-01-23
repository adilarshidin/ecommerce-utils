import pandas as pd
import glob
import os
from openpyxl import load_workbook
from pathlib import Path

# =========================
# PATHS
# =========================
CSV_FILE = "output/all_listings_with_images_and_category.csv"
XLSX_DIR = "templates/worten"
OUTPUT_DIR = "output/worten_filled"

Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

# =========================
# AMAZON → WORTEN MAPPING
# =========================
WORTEN_MAPPING = {
    "moda": {
        "APPAREL", "ONE_PIECE_OUTFIT",
        "COAT", "SHIRT", "PANTS", "SHORTS", "SKIRT",
        "SWEATER", "SWEATSHIRT", "VEST",
        "SOCKS", "UNDERPANTS", "BASE_LAYER_APPAREL_SET",
        "HAT", "SCARF", "EARMUFF",
        "SWIMWEAR", "SNOWSUIT",
        "APPAREL_GLOVES", "SPORT_ACTIVITY_GLOVE",
        "BELTS", "SUSPENDER",
    },

    "ropa_y_calzado_deportivo": {
        "SHOES", "BOOT",
        "SPORT_HELMET", "AUTOMOTIVE_HELMET",
        "SPORTING_GOODS",
        "SPORT_EQUIPMENT_BAG_CASE",
        "KNEE_PAD", "SAFETY_GLASSES",
        "SNOWSHOE",
        "HYDRATION_PACK",
        "SPORT_BAT",
    },

    "salud_bienestar_y_cuidados_para_bebe": {
        "HEALTH_PERSONAL_CARE",
        "FIRST_AID_KIT",
        "SELF_DEFENSE_SPRAY",
        "PROTECTIVE_GLOVE",
        "BODY_PAINT",
        "SAFETY_HARNESS",
    },

    "productos_de_cuidado_personal": {
        "COSMETIC_CASE",
        "BODY_PAINT",
        "HEALTH_PERSONAL_CARE",
    },

    "supermercado_bebidas_y_limpieza": {
        "FOOD",
        "CLEANING_AGENT",
        "PEST_CONTROL_DEVICE",
        "SOLID_FIRE_FUEL",
    },

    "muebles_y_accesorios": {
        "HOME_FURNITURE_AND_DECOR",
        "STOOL_SEATING",
        "TABLE",
        "PILLOW",
        "BLANKET",
        "STORAGE_BOX",
    },

    "deporte_aire_libre_y_viaje": {
        "TENT", "TARP", "HAMMOCK",
        "SLEEPING_BAG", "SLEEPING_MAT",
        "OUTDOOR_RECREATION_PRODUCT",
        "BACKPACK", "DUFFEL_BAG", "CARRIER_BAG_CASE",
        "HYDRATION_PACK",
    },

    "bricolaje_y_construccion": {
        "AXE", "KNIFE", "MULTITOOL", "SCREW_GUN", "SHOVEL_SPADE", "SAW",
    },

    "hogar": {
        "COOKING_POT", "KITCHEN_KNIFE", "FLATWARE", "DISHWARE_PLACE_SETTING", "DISHWARE_PLATE",
        "THERMOS", "DRINKING_CUP", "BOTTLE", "TOWEL", "PILLOW", "BLANKET", "STORAGE_BAG",
    },

    "merchandising_&_gifting": {
        "GLITTER", "CHARM", "BADGE_HOLDER", "BANNER", "KEYCHAIN", "LABEL",
    },

    "smart_home": {
        "LIGHT_BULB", "LIGHT_FIXTURE", "HOME_LIGHTING_AND_LAMPS",
    }
}

# =========================
# LOAD CSV
# =========================
df = pd.read_csv(CSV_FILE, dtype=str)

required_cols = {"seller-sku", "amazon_product_type"}
missing = required_cols - set(df.columns)
if missing:
    raise ValueError(f"Missing required CSV columns: {missing}")

# =========================
# PROCESS EACH XLSX
# =========================
xlsx_files = glob.glob(os.path.join(XLSX_DIR, "*.xlsx"))

for xlsx_path in xlsx_files:
    filename = os.path.splitext(os.path.basename(xlsx_path))[0]

    if filename not in WORTEN_MAPPING:
        print(f"⏭️ Skipping {filename}.xlsx (no mapping)")
        continue

    amazon_types = WORTEN_MAPPING[filename]

    matched_skus = df[
        df["amazon_product_type"].isin(amazon_types)
    ]["seller-sku"].dropna().unique().tolist()

    if not matched_skus:
        print(f"⚠️ No matches for {filename}.xlsx")
        continue

    print(f"📦 Writing {len(matched_skus)} SKUs → {filename}.xlsx")

    wb = load_workbook(xlsx_path)
    if "Data" not in wb.sheetnames:
        print(f"❌ Sheet 'Data' not found in {filename}.xlsx")
        continue

    ws = wb["Data"]

    # Find product_id column (row 2)
    product_id_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=2, column=col).value == "product_id":
            product_id_col = col
            break

    if not product_id_col:
        raise ValueError(f"'product_id' column not found in {filename}.xlsx")

    # Find first empty row (starting from row 3)
    row = 3
    while ws.cell(row=row, column=product_id_col).value:
        row += 1

    # Write SKUs
    for sku in matched_skus:
        ws.cell(row=row, column=product_id_col, value=sku)
        row += 1

    # Save to output directory instead of overwriting input
    output_path = os.path.join(OUTPUT_DIR, os.path.basename(xlsx_path))
    wb.save(output_path)

print(f"✅ All applicable Worten sheets updated successfully in {OUTPUT_DIR}.")
